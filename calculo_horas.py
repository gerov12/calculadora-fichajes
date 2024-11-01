import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog
import os
import sys  # Importamos sys para poder salir del programa

def procesar_fichajes(archivo_txt, archivo_excel):
    # Lista para almacenar los datos
    datos = []

    # Leer el archivo de texto
    with open(archivo_txt, 'r') as file:
        for linea in file:
            linea = linea.encode('ascii', 'ignore').decode()  # Elimina caracteres no ASCII como "Â"
            # Separar la línea en partes usando split sin argumentos
            partes = linea.strip().split()
            if len(partes) >= 4:
                # Extraer la información y recortar espacios innecesarios
                fecha = partes[0].strip()
                hora = partes[1].strip()
                accion = partes[2].strip()
                lugar = partes[3].strip()
                # Agregar los datos a la lista
                datos.append([fecha, hora, accion, lugar])

    # Crear un DataFrame de pandas
    df = pd.DataFrame(datos, columns=['fecha', 'hora', 'acción', 'lugar'])

    # Ordenar el DataFrame por la columna 'fecha' y 'hora'
    df['fecha'] = pd.to_datetime(df['fecha'] + ' ' + df['hora'], format='%d/%m/%Y %H:%M:%S')
    df = df.sort_values(by='fecha')
    # Eliminar duplicados (si los hay)
    df = df.drop_duplicates() 

    # Convertir las fechas al formato correcto para visualización en Excel
    df['fecha'] = df['fecha'].dt.strftime('%d/%m/%Y')

    # Renombrar columnas a mayúscula
    df.columns = [col.capitalize() for col in df.columns]

    # Guardar el DataFrame en un archivo de Excel
    df.to_excel(archivo_excel, index=False, sheet_name='Registros')

    # Calcular horas trabajadas por día
    calcular_horas_por_dia(df, archivo_excel)

    # Calcular horas trabajadas por mes
    calcular_estadisticas_por_mes(df, archivo_excel)

    # Reiniciar el índice para evitar problemas con las filas insertadas
    df = df.reset_index(drop=True)  
    # Agregar líneas separadoras cuando cambia la fecha
    agregar_lineas_separadoras(df, archivo_excel)
    
    print(f"Proceso completado. Los resultados se guardaron en '{archivo_excel}'.")


def calcular_horas_por_dia(df, archivo_excel):
    # Diccionario para almacenar las horas trabajadas por fecha
    horas_trabajadas = {}

    # Inicializar variables
    entrada = None

    for index, row in df.iterrows():
        fecha = row['Fecha']
        hora = row['Hora']
        accion = row['Acción']

        # Convertir la hora a datetime
        datetime_actual = pd.to_datetime(fecha + ' ' + hora, format='%d/%m/%Y %H:%M:%S')

        if accion == "Entrada":
            # Guardar la entrada
            entrada = datetime_actual
        elif accion == "Salida" and entrada is not None:
            # Calcular el tiempo trabajado desde la última entrada
            tiempo_trabajado = datetime_actual - entrada
            horas_trabajadas[fecha] = horas_trabajadas.get(fecha, pd.Timedelta(0)) + tiempo_trabajado
            entrada = None  # Reiniciar la entrada después de calcular

    # Crear un DataFrame para las horas trabajadas
    total_horas_df = pd.DataFrame(horas_trabajadas.items(), columns=['Fecha', 'Total horas'])

    # Ordenar el DataFrame por fecha
    total_horas_df['Fecha'] = pd.to_datetime(total_horas_df['Fecha'], format='%d/%m/%Y')
    total_horas_df = total_horas_df.sort_values(by='Fecha')

    # Convertir las fechas al formato correcto para visualización en Excel
    total_horas_df['Fecha'] = total_horas_df['Fecha'].dt.strftime('%d/%m/%Y')

    # Convertir total a formato hh:mm:ss
    total_horas_df['Total horas'] = total_horas_df['Total horas'].dt.components.apply(
        lambda x: f"{int(x['hours']):02}:{int(x['minutes']):02}:{int(x['seconds']):02}", axis=1)

    # Guardar el DataFrame de totales en una nueva hoja del mismo archivo
    with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a') as writer:
        total_horas_df.to_excel(writer, index=False, sheet_name='Estadísticas por Día')


def calcular_estadisticas_por_mes(df, archivo_excel):
    # Primero, calculamos las horas trabajadas por día como un DataFrame intermedio
    total_horas_df = pd.DataFrame(columns=['Fecha', 'Total horas'])
    
    entrada = None
    horas_trabajadas = {}
    
    for index, row in df.iterrows():
        fecha = row['Fecha']
        hora = row['Hora']
        accion = row['Acción']
        datetime_actual = pd.to_datetime(fecha + ' ' + hora, format='%d/%m/%Y %H:%M:%S')

        if accion == "Entrada":
            entrada = datetime_actual
        elif accion == "Salida" and entrada is not None:
            tiempo_trabajado = datetime_actual - entrada
            horas_trabajadas[fecha] = horas_trabajadas.get(fecha, pd.Timedelta(0)) + tiempo_trabajado
            entrada = None

    total_horas_df = pd.DataFrame(horas_trabajadas.items(), columns=['Fecha', 'Total horas'])

    # Convertir la columna 'Fecha' a datetime
    total_horas_df['Fecha'] = pd.to_datetime(total_horas_df['Fecha'], format='%d/%m/%Y')

    # Extraer el mes y el año como periodo
    total_horas_df['Mes/Año'] = total_horas_df['Fecha'].dt.to_period('M')

    # Convertir 'Total horas' de hh:mm:ss a Timedelta para poder sumar correctamente
    total_horas_df['Total horas'] = pd.to_timedelta(total_horas_df['Total horas'])

    # Agrupar por Mes/Año y calcular las estadísticas
    estadisticas = total_horas_df.groupby('Mes/Año').agg(
        Total_Horas=('Total horas', 'sum'),
        Total_días=('Total horas', 'count'),
    ).reset_index()

    # Calcular el promedio de horas por día
    estadisticas['Promedio'] = estadisticas['Total_Horas'] / estadisticas['Total_días']

    # Convertir Total_Horas y Promedio a formato hh:mm:ss
    estadisticas['Total_Horas'] = estadisticas['Total_Horas'].apply(
        lambda x: f"{int(x.components.hours + x.components.days * 24):02}:{int(x.components.minutes):02}:{int(x.components.seconds):02}"
    )
    estadisticas['Promedio'] = estadisticas['Promedio'].apply(
        lambda x: f"{int(x.components.hours):02}:{int(x.components.minutes):02}:{int(x.components.seconds):02}"
    )

    # Formatear la columna 'Mes/Año' como mm/yyyy para presentación
    estadisticas['Mes/Año'] = estadisticas['Mes/Año'].dt.strftime('%m/%Y')

    # Ordenar el DataFrame por el periodo original de Mes/Año
    estadisticas = estadisticas.sort_values(by='Mes/Año')

    # Renombrar columnas
    estadisticas.columns = ['Mes/Año', 'Total Horas', 'Total días', 'Promedio']

    # Guardar el DataFrame de estadísticas en una nueva hoja
    with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a') as writer:
        estadisticas.to_excel(writer, index=False, sheet_name='Estadísticas por Mes')


def agregar_lineas_separadoras(df, archivo_excel):
    from openpyxl.styles import Border, Side
    from openpyxl import load_workbook

    # Abrir el archivo Excel
    wb = load_workbook(archivo_excel)
    ws = wb['Registros']

    # Definir el estilo de la línea gruesa
    border = Border(top=Side(border_style='thick'))

    # Comenzar desde la fila 2, ya que la 1 es el encabezado
    fila_inicio = 2
    fecha_anterior = None
    for i in range(fila_inicio, len(df) + fila_inicio):
        fecha_actual = ws.cell(row=i, column=1).value  # Obtener la fecha de la columna 1
        if fecha_actual != fecha_anterior:
            if fecha_anterior is not None:  # Si no es la primera fecha, inserta la línea
                for col in range(1, len(df.columns) + 1):  # Aplica el borde a cada columna de la fila insertada
                    ws.cell(row=i, column=col).border = border
        fecha_anterior = fecha_actual

    # Guardar los cambios en el archivo Excel
    wb.save(archivo_excel)


if __name__ == "__main__":
    # Mostrar el mensaje inicial
    print(
        "1- Ingrese a http://gestionrrhh.indec.gob.ar:9090/sse_generico/generico_login.jsp\n"
        "2- Busque los fichajes de interés\n"
        "3- Copie los datos de la tabla resultante y péguelos en un archivo de texto "
        "(si hay varias páginas simplemente pegar los datos unos debajo de los otros)\n"
        "4- Una vez tenga el archivo de texto presione ENTER para proceder a seleccionar el dicho archivo y comenzar el calculo de sus horas.\n\n"
        "Presione ENTER para continuar..."
    )

    # Esperar que el usuario presione ENTER para continuar
    input()

    # Ocultar la ventana de Tkinter
    root = Tk()
    root.withdraw()
    root.call('wm', 'attributes', '.', '-topmost', '1')  # Llevar al frente

    # Abrir cuadro de diálogo para seleccionar el archivo de texto
    archivo_txt = filedialog.askopenfilename(
        title="Seleccione el archivo de texto",
        filetypes=[("Archivos de texto", "*.txt")]
    )

    # Cerrar la ventana de Tkinter después de seleccionar el archivo
    root.destroy()

    # Verificar si se seleccionó un archivo
    if not archivo_txt:
        print("No se seleccionó ningún archivo. Saliendo del programa.")
        sys.exit()

    # Solicitar el nombre del archivo Excel desde la consola
    nombre_excel = input("Ingrese el nombre del archivo Excel (sin extensión): ")

    # Verificar si se ingresó un nombre
    if not nombre_excel:
        print("No se ingresó ningún nombre para el archivo Excel. Saliendo del programa.")
        sys.exit()

    # Crear la ruta para guardar el archivo Excel
    carpeta_resultados = "resultados"
    if not os.path.exists(carpeta_resultados):
        os.makedirs(carpeta_resultados)
    
    archivo_excel = os.path.join(carpeta_resultados, f"{nombre_excel}.xlsx")

    # Ejecutar la función principal
    procesar_fichajes(archivo_txt, archivo_excel)
